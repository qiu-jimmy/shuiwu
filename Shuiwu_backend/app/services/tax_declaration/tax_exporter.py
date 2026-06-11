"""
报税申报表导出服务
支持导出为Excel、PDF等格式
"""
import io
from typing import Dict, Any
from datetime import datetime


class TaxDeclarationExporter:
    """报税申报表导出器"""

    def export_to_excel(
        self,
        declaration_data: Dict[str, Any],
        format_type: str = "xlsx"
    ) -> bytes:
        """
        导出为Excel格式

        Args:
            declaration_data: 申报数据
            format_type: 格式类型 (xlsx/xls)

        Returns:
            文件内容（bytes）
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "报税申报表"

            # 定义样式
            header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            cell_alignment = Alignment(horizontal='left', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20

            # 标题
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = '报税申报表'
            title_cell.font = Font(name='微软雅黑', size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 申报单号
            row = 3
            ws[f'A{row}'] = '申报单号'
            ws[f'B{row}'] = declaration_data.get('declaration_no', '')
            ws[f'C{row}'] = ''

            # 基本信息
            row += 2
            ws[f'A{row}'] = '纳税人信息'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:C{row}')

            row += 1
            fields = [
                ('纳税人姓名', declaration_data.get('taxpayer_name')),
                ('联系电话', declaration_data.get('taxpayer_phone')),
                ('身份证号', declaration_data.get('taxpayer_id_card')),
                ('纳税人类型', declaration_data.get('taxpayer_type')),
                ('税种', declaration_data.get('tax_type')),
                ('税期', declaration_data.get('tax_period')),
            ]

            for label, value in fields:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value or ''
                ws[f'C{row}'] = ''
                # 应用样式
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.border = border
                    cell.alignment = cell_alignment
                row += 1

            # 收入信息
            row += 1
            ws[f'A{row}'] = '收入信息'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:C{row}')

            row += 1
            income_info = declaration_data.get('income_info') or {}
            for key, value in income_info.items():
                ws[f'A{row}'] = self._translate_field(key)
                ws[f'B{row}'] = value
                ws[f'C{row}'] = '元'
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.border = border
                    cell.alignment = cell_alignment
                row += 1

            # 扣除信息
            row += 1
            ws[f'A{row}'] = '扣除信息'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:C{row}')

            row += 1
            deduction_info = declaration_data.get('deduction_info') or {}
            for key, value in deduction_info.items():
                ws[f'A{row}'] = self._translate_field(key)
                ws[f'B{row}'] = value
                ws[f'C{row}'] = '元'
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.border = border
                    cell.alignment = cell_alignment
                row += 1

            # 计算结果
            if declaration_data.get('total_income') is not None:
                row += 1
                ws[f'A{row}'] = '计算结果'
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].alignment = header_alignment
                ws.merge_cells(f'A{row}:C{row}')

                row += 1
                calc_fields = [
                    ('收入总额', declaration_data.get('total_income')),
                    ('扣除总额', declaration_data.get('total_deduction')),
                    ('应纳税所得额', declaration_data.get('taxable_income')),
                    ('应纳税额', declaration_data.get('tax_amount')),
                    ('已缴税额', declaration_data.get('tax_paid')),
                    ('应退税额', declaration_data.get('tax_refund')),
                ]

                for label, value in calc_fields:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = f'{value:.2f}' if value is not None else ''
                    ws[f'C{row}'] = '元'
                    for col in ['A', 'B', 'C']:
                        cell = ws[f'{col}{row}']
                        cell.border = border
                        cell.alignment = cell_alignment
                    row += 1

            # 状态信息
            row += 1
            ws[f'A{row}'] = '处理状态'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:C{row}')

            row += 1
            status_fields = [
                ('状态', declaration_data.get('status')),
                ('申报流水号', declaration_data.get('declaration_serial_no')),
                ('申报日期', declaration_data.get('declaration_date')),
                ('处理结果', declaration_data.get('process_result')),
                ('处理备注', declaration_data.get('process_notes')),
            ]

            for label, value in status_fields:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value or ''
                ws[f'C{row}'] = ''
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.border = border
                    cell.alignment = cell_alignment
                row += 1

            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output.read()

        except ImportError:
            # 如果没有安装 openpyxl，返回CSV格式
            return self.export_to_csv(declaration_data)

    def export_to_csv(self, declaration_data: Dict[str, Any]) -> bytes:
        """导出为CSV格式"""
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        # 写入标题
        writer.writerow(['报税申报表'])
        writer.writerow(['申报单号', declaration_data.get('declaration_no', '')])
        writer.writerow([])

        # 基本信息
        writer.writerow(['纳税人信息'])
        fields = [
            ['纳税人姓名', declaration_data.get('taxpayer_name')],
            ['联系电话', declaration_data.get('taxpayer_phone')],
            ['身份证号', declaration_data.get('taxpayer_id_card')],
            ['税种', declaration_data.get('tax_type')],
            ['税期', declaration_data.get('tax_period')],
        ]
        writer.writerows(fields)
        writer.writerow([])

        # 计算结果
        if declaration_data.get('total_income'):
            writer.writerow(['计算结果'])
            calc_fields = [
                ['收入总额', declaration_data.get('total_income')],
                ['应纳税额', declaration_data.get('tax_amount')],
            ]
            writer.writerows(calc_fields)

        output.seek(0)
        return output.getvalue().encode('utf-8-sig')  # UTF-8 with BOM for Excel

    def _translate_field(self, field: str) -> str:
        """翻译字段名为中文"""
        translations = {
            'salary': '工资薪金',
            'bonus': '奖金',
            'labor_income': '劳务报酬',
            'author_income': '稿酬收入',
            'royalty_income': '特许权使用费',
            'special_deduction': '专项扣除',
            'additional_deduction': '专项附加扣除',
            'children_education': '子女教育',
            'continuing_education': '继续教育',
            'housing_loan': '住房贷款利息',
            'housing_rent': '住房租金',
            'elderly_support': '赡养老人',
            'infant_care': '婴幼儿照护',
        }
        return translations.get(field, field)


# 全局导出器实例
tax_exporter = TaxDeclarationExporter()
