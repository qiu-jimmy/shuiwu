"""
自定义 Excel 文件读取器

解决 agno CSVReader 处理 xlsx 文件时的编码问题。
xlsx 是二进制格式（ZIP 压缩），不能使用文本模式（utf-8）读取。

使用轻量级的 openpyxl 库，避免引入庞大的 pandas 依赖。
"""
import io
import csv
import asyncio


class ExcelReader:
    """Excel 文件读取器，使用 openpyxl 处理 .xlsx 文件"""

    def __init__(self, chunking_strategy=None):
        """
        初始化 Excel Reader

        Args:
            chunking_strategy: 分块策略（与 agno 兼容）
        """
        self.chunking_strategy = chunking_strategy

    def read(self, file_path_or_buffer, name: str = None, **kwargs) -> list:
        """
        读取 Excel 文件

        Args:
            file_path_or_buffer: 文件路径（字符串或Path对象）、文件对象或字节流
            name: 可选的文件名（与 agno 兼容，实际使用文件名从路径提取）
            **kwargs: 其他参数（与 agno 兼容）

        Returns:
            Document 对象列表
        """
        from agno.knowledge.document import Document
        from pathlib import Path

        # 读取文件内容为字节流
        if isinstance(file_path_or_buffer, (str, Path)):
            # 文件路径（字符串或 Path 对象）
            path = Path(file_path_or_buffer) if isinstance(file_path_or_buffer, str) else file_path_or_buffer
            with open(path, 'rb') as f:
                content = f.read()
        elif isinstance(file_path_or_buffer, bytes):
            # 字节流
            content = file_path_or_buffer
        else:
            # 文件对象
            content = file_path_or_buffer.read()

        # 使用 openpyxl 读取所有工作表
        sheets_data = self._read_excel_with_openpyxl(content)

        # 转换为 Document 列表
        from agno.knowledge.document import Document

        documents = []
        for sheet_name, text_content in sheets_data.items():
            # 先创建完整的 Document
            doc = Document(
                content=text_content,
                name=sheet_name,
                meta_data={"sheet": sheet_name}
            )

            # 应用分块策略
            if self.chunking_strategy:
                try:
                    chunked_docs = self.chunking_strategy.chunk(doc)
                    documents.extend(chunked_docs)
                except Exception as e:
                    # 如果分块失败，使用原始文档
                    print(f"分块失败，使用原始文档: {e}")
                    documents.append(doc)
            else:
                documents.append(doc)

        return documents

    async def async_read(self, file_path_or_buffer, name: str = None) -> list:
        """
        异步读取 Excel 文件

        Args:
            file_path_or_buffer: 文件路径、文件对象或字节流
            name: 可选的文件名（用于命名文档）

        Returns:
            Document 对象列表
        """
        # Excel 读取本身是同步的（openpyxl 不支持异步），
        # 所以我们在线程池中运行同步读取，避免阻塞事件循环
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self.read, file_path_or_buffer)

    def _read_excel_with_openpyxl(self, file_content: bytes) -> dict:
        """
        使用 openpyxl 读取 Excel 文件的所有工作表

        Args:
            file_content: 文件字节内容

        Returns:
            字典，key 为 sheet 名，value 为格式化的文本内容
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        # 从字节流加载工作簿
        buffer = io.BytesIO(file_content)

        try:
            wb = load_workbook(filename=buffer, read_only=True, data_only=True)
        except Exception as e:
            # 如果 openpyxl 失败（可能是旧版 .xls 文件），尝试使用 xlrd
            if "not a zip file" in str(e).lower() or "bad zip file" in str(e).lower():
                print(f"[WARN] openpyxl 读取失败，尝试使用 xlrd: {e}")
                return self._read_excel_with_xlrd(file_content)
            raise

        sheets_data = {}

        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_content = self._sheet_to_text(ws, sheet_name)
            sheets_data[sheet_name] = text_content

        wb.close()
        return sheets_data

    def _read_excel_with_xlrd(self, file_content: bytes) -> dict:
        """
        使用 xlrd 读取旧版 Excel 文件（.xls 格式）

        Args:
            file_content: 文件字节内容

        Returns:
            字典，key 为 sheet 名，value 为格式化的文本内容
        """
        try:
            import xlrd
        except ImportError:
            raise ImportError("读取旧版 .xls 文件需要安装 xlrd: pip install xlrd")

        # 从字节流加载工作簿
        buffer = io.BytesIO(file_content)
        wb = xlrd.open_workbook(file_contents=buffer.read())

        sheets_data = {}

        # 遍历所有工作表
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            text_content = self._xls_sheet_to_text(sheet, sheet_name)
            sheets_data[sheet_name] = text_content

        return sheets_data

    def _xls_sheet_to_text(self, worksheet, sheet_name: str) -> str:
        """
        将 xlrd 工作表转换为文本格式

        Args:
            worksheet: xlrd 工作表对象
            sheet_name: 工作表名称

        Returns:
            格式化的文本
        """
        lines = [f"## 工作表: {sheet_name}"]

        if worksheet.nrows == 0:
            lines.append("(空工作表)")
            return "\n".join(lines)

        # 处理表头（第一行）
        headers = [worksheet.cell_value(0, col) for col in range(worksheet.ncols)]
        header_str = " | ".join([str(cell) if cell else "" for cell in headers])
        lines.append(header_str)
        lines.append("-" * len(header_str))

        # 处理数据行
        for row_idx in range(1, worksheet.nrows):
            row = [worksheet.cell_value(row_idx, col) for col in range(worksheet.ncols)]
            # 跳过完全空白的行
            if all(not cell or str(cell).strip() == "" for cell in row):
                continue
            row_str = " | ".join([str(cell) if cell else "" for cell in row])
            lines.append(row_str)

        # 添加统计信息
        non_empty_rows = sum(
            1 for row_idx in range(1, worksheet.nrows)
            if any(worksheet.cell_value(row_idx, col) for col in range(worksheet.ncols))
        )
        lines.append(f"\n数据行数: {non_empty_rows}, 列数: {worksheet.ncols}")

        return "\n".join(lines)

    def _sheet_to_text(self, worksheet, sheet_name: str) -> str:
        """
        将工作表转换为文本格式

        Args:
            worksheet: openpyxl 工作表对象
            sheet_name: 工作表名称

        Returns:
            格式化的文本
        """
        lines = [f"## 工作表: {sheet_name}"]

        # 获取所有数据行
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            lines.append("(空工作表)")
            return "\n".join(lines)

        # 处理表头（第一行）
        headers = rows[0]
        header_str = " | ".join([str(cell) if cell is not None else "" for cell in headers])
        lines.append(header_str)
        lines.append("-" * len(header_str))

        # 处理数据行
        data_rows = rows[1:] if len(rows) > 1 else []
        for row in data_rows:
            # 跳过完全空白的行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
            lines.append(row_str)

        # 添加统计信息
        non_empty_rows = sum(
            1 for row in data_rows
            if any(cell is not None and str(cell).strip() != "" for cell in row)
        )
        lines.append(f"\n数据行数: {non_empty_rows}, 列数: {len(headers)}")

        return "\n".join(lines)


def create_excel_reader(chunking_strategy=None) -> ExcelReader:
    """
    创建 Excel Reader 实例的工厂函数

    Args:
        chunking_strategy: 分块策略

    Returns:
        ExcelReader 实例
    """
    return ExcelReader(chunking_strategy=chunking_strategy)
